from __future__ import annotations

from dataclasses import asdict
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from warframe_app.clients import WarframeMarketClient, WorldstateClient
from warframe_app.models import (
    ActiveEvent,
    ActivityItem,
    AlertItem,
    BaroDeal,
    DashboardSnapshot,
    FeatureModule,
    FissureMission,
    IntegrationNeed,
    LiveCycle,
    MarketCatalogItem,
    MarketLiveOrder,
    MarketMetric,
    MarketSnapshot,
    MarketTradeStat,
    MissingPiece,
    NextStep,
    RelicDrop,
    RelicSnapshot,
    VendorStatus,
    VosforSnapshot,
    VosforSourceEntry,
    VosforTopEntry,
    WorldstateSnapshot,
)
from warframe_app.storage import AppStorage


class WarframeDataService:
    def __init__(self) -> None:
        self.storage = AppStorage()
        self.worldstate_client = WorldstateClient(platform="pc")
        self.market_client = WarframeMarketClient()
        self.vosfor_workbook_path = Path.home() / "Downloads" / "Vosfor Gambling.xlsx"
        self._market_seed_items = [
            MarketCatalogItem("Arcane Energize", "arcane_energize"),
            MarketCatalogItem("Arcane Barrier", "arcane_barrier"),
            MarketCatalogItem("Glaive Prime Set", "glaive_prime_set", is_set=True),
            MarketCatalogItem("Revenant Prime Set", "revenant_prime_set", is_set=True),
        ]
        self._activities = [
            ActivityItem(
                name="Phase 1: Live Data Backbone",
                focus="Worldstate + Timers",
                note="Reset timer, open world cycles, fissures, events, baro, bounties und notifications zuerst.",
            ),
            ActivityItem(
                name="Phase 2: Account Progression",
                focus="Inventory + Foundry",
                note="Foundry, inventory, mastery helper, material planner und craft tracking auf denselben Datensatz setzen.",
            ),
            ActivityItem(
                name="Phase 3: Trade Intelligence",
                focus="Market + Relics",
                note="Smart alerts, fake-list filter, relic planner, analyzer, rivens und quick sell actions bauen.",
            ),
            ActivityItem(
                name="Phase 4: Community + Overlay",
                focus="Overlay + Sharing",
                note="While-you-were-away, party tools, clan search, builds und Overlay nach dem Desktop-Kern.",
            ),
        ]
        self._feature_modules = [
            FeatureModule(
                phase="1",
                area="Core Live Data",
                name="Reset Timers + Rotations",
                summary="Alle globalen Reset-Timer, open world cycles, NPC-Rotationen und Daily/Weekly-Resets zentral anzeigen.",
                dependencies="Worldstate API, lokale Notification-Engine, Plattformwahl",
                release="In progress",
            ),
            FeatureModule(
                phase="1",
                area="Core Live Data",
                name="Void Fissure Tracker",
                summary="Normale und Steel Path Fissures mit Timern, Filtern und Benachrichtigungen.",
                dependencies="Worldstate API fuer fissures, Notification-Engine",
                release="In progress",
            ),
            FeatureModule(
                phase="1",
                area="Core Live Data",
                name="Events + NPC Rotation",
                summary="Events, Baro Ki'Teer, Varzia, Tenet-Angebote und NPC-Shops inklusive aktueller Waren.",
                dependencies="Worldstate API, static item catalog",
                release="In progress",
            ),
            FeatureModule(
                phase="1",
                area="Economy",
                name="Vosfor Analyzer",
                summary="Importiert dein eigenes Vosfor-Sheet und zeigt Gambling-, Baro- und Plat-Effizienz in der App.",
                dependencies="Lokaler Excel-Import, SQLite",
                release="In progress",
            ),
            FeatureModule(
                phase="2",
                area="Account Progression",
                name="Foundry Overview",
                summary="AlecaFrame-aehnliche Foundry mit craftbaren, fertigen und vorbereiteten Items.",
                dependencies="Inventarzugriff, blueprint/item database, lokaler Speicher",
                release="Planned",
            ),
            FeatureModule(
                phase="2",
                area="Account Progression",
                name="Mastery Helper",
                summary="Offene MR-Items, Mastery-Pfade und Mass-Mastery-Plan mit gemeinsamen Farmspots.",
                dependencies="Inventarzugriff, blueprint graph, mastery metadata",
                release="Planned",
            ),
            FeatureModule(
                phase="2",
                area="Account Progression",
                name="Inventory + Drop Sources",
                summary="Inventar mit fehlenden Komponenten, Drop-Locations, Kaufquellen und Lich/Status-Hinweisen.",
                dependencies="Inventarzugriff, static drop tables, item metadata",
                release="Planned",
            ),
            FeatureModule(
                phase="3",
                area="Trade Intelligence",
                name="Warframe Market Panel",
                summary="Marktpanel mit ausblendbarer Sidebar, Crossplay-Filtern und Average-Price-Daten.",
                dependencies="warframe.market public API, authentifizierter Session-Flow fuer eigene Orders",
                release="In progress",
            ),
            FeatureModule(
                phase="3",
                area="Trade Intelligence",
                name="Smart Alerts + Fake List Filter",
                summary="Preisalarme, Filter gegen offensichtliche Fake-Listings und Event-Impact-Hinweise wie Prime Resurgence.",
                dependencies="warframe.market stats, Preisregeln, lokaler Alert-Speicher",
                release="Planned",
            ),
            FeatureModule(
                phase="3",
                area="Trade Intelligence",
                name="Relic Planner 2.0",
                summary="Relics nach Plat-, Ducat- und Inventory-Wert planen und geoeffnete Relics aus dem Pool entfernen.",
                dependencies="Relic inventory data, market stats, local run history",
                release="Planned",
            ),
            FeatureModule(
                phase="4",
                area="Community + Overlay",
                name="Overlay + Sharing",
                summary="Build-Sharing, Overlay-Anpassung, Party-System und Clan-Recruiting nach dem Desktop-Kern.",
                dependencies="Eigener Backend-Service, OCR stack, draggable overlay window",
                release="Later",
            ),
        ]
        self._integrations = [
            IntegrationNeed(
                layer="Live Worldstate",
                source="warframestat.us API",
                access="Public HTTP, kein API-Key",
                used_for="Fissures, cycles, syndicate/bounty rotations, events, Baro, arbitration und globale Timer",
                priority="Implemented",
            ),
            IntegrationNeed(
                layer="Workbook Import",
                source="Lokales Excel-Sheet",
                access="Lesender Dateizugriff",
                used_for="Vosfor Gambling, Baro-Profit und eigene Wirtschaftstabellen",
                priority="Implemented",
            ),
            IntegrationNeed(
                layer="Persistence",
                source="Lokale SQLite-Datenbank",
                access="App-intern",
                used_for="API-Cache, importierte Vosfor-Daten, Alerts und spaeter Nutzerzustand",
                priority="Implemented",
            ),
            IntegrationNeed(
                layer="Market Data",
                source="warframe.market API",
                access="Live ueber Browser-Session fuer Orders und v2-Endpoints, spaeter Auth fuer eigene Listings",
                used_for="Ingame-Orders, Seller-Namen, Whisper-Text, Bilder, Closed-Trade-Stats und spaeter eigene Orders",
                priority="Live ingame orders implemented",
            ),
            IntegrationNeed(
                layer="Inventory Sync",
                source="Overwolf-artige Inventarquelle oder Alternative",
                access="Nicht oeffentlich trivial; braucht lokale Game-/Overlay-Integration oder manuellen Import",
                used_for="Foundry, Mastery, Inventory-Tracking, Relic Counts, Crafting und Sell/Keep Analyzer",
                priority="Critical blocker",
            ),
        ]
        self._missing_pieces = [
            MissingPiece(
                area="Inventarzugriff",
                impact="Ohne ihn koennen Foundry, Mastery Helper, Inventory-Tracking und Relic Counts nicht echt funktionieren.",
                needed_from_user="Entscheidung: Overwolf-aehnliche Integration, manueller Export oder OCR/Log-basierter Hybrid",
                notes="Das ist weiter der groesste technische Blocker.",
            ),
            MissingPiece(
                area="warframe.market Write Access",
                impact="Read-only reicht fuer Preise, aber Bulk-Actions und eigene Orders brauchen Login/Session-Handling.",
                needed_from_user="Spaeter Login-Strategie festlegen",
                notes="Read-only Live-Orders laufen, aber eigene Listings und Bulk-Aktionen brauchen weiter Login/Session-Handling.",
            ),
            MissingPiece(
                area="Drop-Table Wahrheit",
                impact="Du willst bessere Drop-Locations als AlecaFrame. Dafuer brauchen wir ein sauberes normalisiertes Datenmodell.",
                needed_from_user="Quelle festlegen oder eigenes Dataset zusammenbauen",
                notes="Mehr Produktarbeit als reine API-Verkabelung.",
            ),
            MissingPiece(
                area="Overlay Strategy",
                impact="Ein frei verschiebbares Overlay braucht extra Fenstermanagement, Hotkeys und Capture-Logik.",
                needed_from_user="Desktop-App, Sidecar-Overlay oder vollwertiges Ingame-Overlay festlegen",
                notes="Wuerde ich nach dem stabilen Desktop-Kern bauen.",
            ),
        ]
        self._next_steps = [
            NextStep(
                step="1. Preisalarme und Watchlist bauen",
                outcome="Die Live-Orders und Stats koennen direkt fuer Alerts, Trends und den spaeteren Sell/Keep-Analyzer genutzt werden.",
            ),
            NextStep(
                step="2. Notifications ergaenzen",
                outcome="Fissures, Events und Reset-Timer koennen lokal Benachrichtigungen ausloesen.",
            ),
            NextStep(
                step="3. Inventarstrategie festlegen",
                outcome="Danach koennen Foundry, Mastery und Inventory wirklich gebaut statt nur vorbereitet werden.",
            ),
        ]
        self._relics = {
            "Axi A18": [
                RelicDrop("Axi A18", "Rare", "Akarius Prime Blueprint", "10%"),
                RelicDrop("Axi A18", "Uncommon", "Cedo Prime Receiver", "20%"),
                RelicDrop("Axi A18", "Common", "Forma Blueprint", "25.33%"),
            ],
            "Neo S19": [
                RelicDrop("Neo S19", "Rare", "Soma Prime Stock", "10%"),
                RelicDrop("Neo S19", "Uncommon", "Sevagoth Prime Chassis", "20%"),
                RelicDrop("Neo S19", "Common", "Paris Prime Lower Limb", "25.33%"),
            ],
            "Lith R14": [
                RelicDrop("Lith R14", "Rare", "Revenant Prime Neuroptics", "10%"),
                RelicDrop("Lith R14", "Uncommon", "Rhino Prime Blueprint", "20%"),
                RelicDrop("Lith R14", "Common", "Bronco Prime Barrel", "25.33%"),
            ],
        }

    def load_dashboard(self) -> DashboardSnapshot:
        live_ops = self.load_worldstate()
        vosfor = self.load_vosfor()
        return DashboardSnapshot(
            last_updated=datetime.now().strftime("%d.%m.%Y %H:%M"),
            alerts=self._build_alerts(live_ops, vosfor),
            activities=self._activities,
            tracked_items=self.list_market_items(),
            tracked_relics=sorted(self._relics.keys()),
            feature_modules=self._feature_modules,
            integrations=self._integrations,
            missing_pieces=self._missing_pieces,
            live_ops=live_ops,
            vosfor=vosfor,
        )

    def load_worldstate(self) -> WorldstateSnapshot:
        cache_key = "worldstate:pc"
        cached_payload, cached_at = self.storage.get_cached_json(cache_key, max_age_seconds=300)
        status = "Live data loaded from SQLite cache."
        payload = cached_payload
        fetched_at = cached_at or datetime.now(UTC).isoformat()

        if payload is None:
            try:
                payload = self.worldstate_client.fetch_worldstate()
                fetched_at = self.storage.set_cached_json(cache_key, payload)
                status = "Live data fetched from warframestat.us."
            except RuntimeError as error:
                payload, fetched_at = self.storage.get_any_cached_json(cache_key)
                if payload is not None:
                    status = f"{error} Offline fallback uses cached data."
                else:
                    return WorldstateSnapshot(
                        fetched_at=self._friendly_timestamp(None),
                        status=str(error),
                        cycles=[],
                        fissures=[],
                        events=[],
                        vendors=[],
                    )

        return WorldstateSnapshot(
            fetched_at=self._friendly_timestamp(fetched_at),
            status=status,
            cycles=self._parse_cycles(payload),
            fissures=self._parse_fissures(payload),
            events=self._parse_events(payload),
            vendors=self._parse_vendors(payload),
        )

    def load_vosfor(self) -> VosforSnapshot:
        if self.vosfor_workbook_path.exists():
            snapshot = self._parse_vosfor_workbook(self.vosfor_workbook_path)
            self.storage.save_state("vosfor_snapshot", self._serialize_vosfor(snapshot))
            return snapshot

        stored = self.storage.load_state("vosfor_snapshot")
        if stored is not None:
            return self._deserialize_vosfor(stored)

        return VosforSnapshot(
            imported_at="Nicht importiert",
            workbook_path=str(self.vosfor_workbook_path),
            gambling_entries=[],
            top_entries=[],
            baro_deals=[],
        )

    def list_feature_modules(self) -> list[FeatureModule]:
        return list(self._feature_modules)

    def list_integrations(self) -> list[IntegrationNeed]:
        return list(self._integrations)

    def list_missing_pieces(self) -> list[MissingPiece]:
        return list(self._missing_pieces)

    def list_next_steps(self) -> list[NextStep]:
        return list(self._next_steps)

    def close(self) -> None:
        self.market_client.close()

    def search_market(self, query: str, force_refresh: bool = False) -> MarketSnapshot:
        catalog = self._load_market_catalog()
        item = self._match_market_item(query, catalog)

        try:
            details_payload, details_at, details_status = self._load_market_details(
                item.url_name,
                force_refresh=force_refresh,
            )
        except RuntimeError as error:
            return self._empty_market_snapshot(item, str(error))

        try:
            orders_payload, orders_at, orders_status = self._load_market_orders(
                item.url_name,
                force_refresh=force_refresh,
            )
        except RuntimeError as error:
            orders_payload, orders_at, orders_status = {"data": []}, None, str(error)

        statistics_payload, statistics_at, statistics_status = self._load_market_statistics(
            item.url_name,
            force_refresh=force_refresh,
        )

        payload = {
            "item_name": item.item_name,
            "url_name": item.url_name,
            "details": details_payload,
            "orders": orders_payload,
            "statistics": statistics_payload,
        }
        status = " | ".join(
            part
            for part in (orders_status, details_status, statistics_status)
            if part
        )
        fetched_at = orders_at or details_at or statistics_at or datetime.now(UTC).isoformat()
        return self._build_market_snapshot(item, payload, fetched_at, status)

    def list_market_items(self) -> list[str]:
        return [item.item_name for item in self._market_seed_items]

    def list_relics(self) -> list[str]:
        return sorted(self._relics.keys())

    def search_relic(self, query: str) -> RelicSnapshot:
        relic_name = self._closest_key(query, self._relics.keys())
        return RelicSnapshot(relic_name=relic_name, drops=self._relics[relic_name])

    def _build_alerts(self, live_ops: WorldstateSnapshot, vosfor: VosforSnapshot) -> list[AlertItem]:
        alerts: list[AlertItem] = []
        steel_path = [fissure for fissure in live_ops.fissures if fissure.is_steel_path]
        if steel_path:
            alerts.append(
                AlertItem(
                    title="Steel Path fissures live",
                    description=f"{len(steel_path)} Steel Path fissures sind gerade aktiv.",
                    severity="high",
                )
            )
        if live_ops.events:
            first_event = live_ops.events[0]
            alerts.append(
                AlertItem(
                    title="Aktives Event",
                    description=f"{first_event.name} auf {first_event.node} ({first_event.eta}).",
                    severity="medium",
                )
            )
        if live_ops.vendors:
            vendor = live_ops.vendors[0]
            alerts.append(
                AlertItem(
                    title=f"{vendor.name} Rotation",
                    description=f"{vendor.location} | {vendor.eta}",
                    severity="info",
                )
            )
        if vosfor.top_entries:
            best = vosfor.top_entries[0]
            alerts.append(
                AlertItem(
                    title="Vosfor Top Entry",
                    description=f"{best.item_name} bringt laut Sheet {best.vosfor_per_plat} Vosfor pro Plat.",
                    severity="info",
                )
            )
        if not alerts:
            alerts.append(
                AlertItem(
                    title="Noch keine Live-Daten",
                    description="Sobald Worldstate oder Vosfor geladen sind, erscheinen hier Prioritaeten.",
                    severity="low",
                )
            )
        return alerts

    def _parse_cycles(self, payload: dict[str, Any]) -> list[LiveCycle]:
        cycle_specs = [
            ("Earth", payload.get("earthCycle")),
            ("Cetus", payload.get("cetusCycle")),
            ("Orb Vallis", payload.get("vallisCycle")),
            ("Cambion Drift", payload.get("cambionCycle")),
            ("Duviri", payload.get("duviriCycle")),
        ]
        cycles: list[LiveCycle] = []
        for name, cycle_payload in cycle_specs:
            cycle = self._build_cycle(name, cycle_payload)
            if cycle is not None:
                cycles.append(cycle)
        return cycles

    def _build_cycle(self, name: str, cycle_payload: Any) -> LiveCycle | None:
        if not isinstance(cycle_payload, dict):
            return None

        state = cycle_payload.get("state")
        if not state:
            if "isDay" in cycle_payload:
                state = "day" if cycle_payload["isDay"] else "night"
            elif "isWarm" in cycle_payload:
                state = "warm" if cycle_payload["isWarm"] else "cold"
            elif "active" in cycle_payload:
                state = "active" if cycle_payload["active"] else "inactive"
            else:
                state = "unknown"

        remaining = (
            cycle_payload.get("timeLeft")
            or cycle_payload.get("shortString")
            or cycle_payload.get("remaining")
            or cycle_payload.get("eta")
            or self._remaining_from_expiry(cycle_payload.get("expiry"))
        )
        remaining = self._format_timer_text(remaining)

        expires_at = self._friendly_timestamp(cycle_payload.get("expiry"))
        if remaining == "expired":
            return None
        return LiveCycle(
            name=name,
            state=str(state),
            remaining=str(remaining),
            expires_at=expires_at,
        )

    def _parse_fissures(self, payload: dict[str, Any]) -> list[FissureMission]:
        fissures: list[FissureMission] = []
        for entry in payload.get("fissures", []):
            if entry.get("expired"):
                continue
            eta = self._format_timer_text(
                entry.get("eta") or self._remaining_from_expiry(entry.get("expiry"))
            )
            if eta == "expired":
                continue
            fissures.append(
                FissureMission(
                    tier=str(entry.get("tier") or entry.get("tierNum") or "Unknown"),
                    mission_type=str(entry.get("missionType") or "Unknown"),
                    node=str(entry.get("node") or "Unknown"),
                    enemy=str(entry.get("enemy") or entry.get("enemyKey") or "-"),
                    eta=eta,
                    is_steel_path=bool(entry.get("isHard")),
                )
            )
        return fissures[:18]

    def _parse_events(self, payload: dict[str, Any]) -> list[ActiveEvent]:
        events: list[ActiveEvent] = []
        for entry in payload.get("events", []):
            rewards: list[str] = []
            for reward in entry.get("rewards", []):
                rewards.extend(reward.get("items", []))
            reward_text = ", ".join(rewards[:3]) if rewards else "-"
            events.append(
                ActiveEvent(
                    name=str(entry.get("description") or entry.get("node") or "Event"),
                    node=str(entry.get("node") or "-"),
                    eta=self._remaining_from_expiry(entry.get("expiry")),
                    rewards=reward_text,
                )
            )
        return events[:12]

    def _parse_vendors(self, payload: dict[str, Any]) -> list[VendorStatus]:
        vendors: list[VendorStatus] = []
        vendor_specs = [
            ("Baro Ki'Teer", payload.get("voidTrader")),
            ("Varzia", payload.get("vaultTrader")),
        ]
        for name, vendor_payload in vendor_specs:
            vendor = self._build_vendor(name, vendor_payload)
            if vendor is not None:
                vendors.append(vendor)
        return vendors

    def _build_vendor(self, name: str, vendor_payload: Any) -> VendorStatus | None:
        if not isinstance(vendor_payload, dict):
            return None
        inventory = vendor_payload.get("inventory", [])
        preview = ", ".join(
            str(item.get("item", "-"))
            for item in inventory[:3]
            if isinstance(item, dict)
        ) or "-"

        eta = self._window_status(
            activation=vendor_payload.get("activation"),
            expiry=vendor_payload.get("expiry"),
        )
        return VendorStatus(
            name=name,
            location=str(vendor_payload.get("location") or "-"),
            eta=eta,
            inventory_preview=preview,
        )

    def _parse_vosfor_workbook(self, workbook_path: Path) -> VosforSnapshot:
        workbook = load_workbook(workbook_path, data_only=True)
        gambling_entries: list[VosforSourceEntry] = []
        top_entries: list[VosforTopEntry] = []
        baro_deals: list[BaroDeal] = []

        if "Vosfor Gambling" in workbook.sheetnames:
            worksheet = workbook["Vosfor Gambling"]
            current_source = "Unknown"
            for row in worksheet.iter_rows(values_only=True):
                first = row[0] if len(row) > 0 else None
                second = row[1] if len(row) > 1 else None
                if first and second == "Chance":
                    current_source = str(first)
                    continue
                if not first or second in (None, "Chance"):
                    continue
                gambling_entries.append(
                    VosforSourceEntry(
                        source=current_source,
                        item_name=str(first),
                        chance=self._format_number(second),
                        avg_price_rank0=self._format_mixed(row[2] if len(row) > 2 else None),
                        avg_price_maxed=self._format_mixed(row[3] if len(row) > 3 else None),
                        expected_rank0=self._format_mixed(row[7] if len(row) > 7 else None),
                        expected_maxed=self._format_mixed(row[8] if len(row) > 8 else None),
                        unit_price_maxed=self._format_mixed(row[9] if len(row) > 9 else None),
                    )
                )

        if "Vosforplat" in workbook.sheetnames:
            worksheet = workbook["Vosforplat"]
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    top_entries.append(
                        VosforTopEntry(
                            item_name=str(row[0]),
                            vosfor_amount=self._format_mixed(row[1]),
                            platinum=self._format_mixed(row[2]),
                            vosfor_per_plat=self._format_mixed(row[3]),
                        )
                    )

        if "Baro kiteer" in workbook.sheetnames:
            worksheet = workbook["Baro kiteer"]
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    baro_deals.append(
                        BaroDeal(
                            item_name=str(row[0]),
                            ducats=self._format_mixed(row[1]),
                            platinum=self._format_mixed(row[2]),
                            profit=self._format_mixed(row[3]),
                        )
                    )

        snapshot = VosforSnapshot(
            imported_at=datetime.now().strftime("%d.%m.%Y %H:%M"),
            workbook_path=str(workbook_path),
            gambling_entries=gambling_entries,
            top_entries=sorted(
                top_entries,
                key=lambda entry: self._safe_float(entry.vosfor_per_plat),
                reverse=True,
            ),
            baro_deals=sorted(
                baro_deals,
                key=lambda entry: self._safe_float(entry.profit),
                reverse=True,
            ),
        )
        workbook.close()
        return snapshot

    def _load_market_catalog(self) -> list[MarketCatalogItem]:
        cache_key = "market:catalog"
        cached_payload, _ = self.storage.get_cached_json(cache_key, max_age_seconds=43200)
        if cached_payload is not None:
            return self._deserialize_market_catalog(cached_payload)

        try:
            items = self.market_client.fetch_item_catalog()
            payload = {"items": items}
            self.storage.set_cached_json(cache_key, payload)
            return self._deserialize_market_catalog(payload)
        except RuntimeError:
            cached_payload, _ = self.storage.get_any_cached_json(cache_key)
            if cached_payload is not None:
                return self._deserialize_market_catalog(cached_payload)
            return list(self._market_seed_items)

    def _deserialize_market_catalog(self, payload: dict[str, Any]) -> list[MarketCatalogItem]:
        items = payload.get("items", [])
        catalog = [
            MarketCatalogItem(
                item_name=str(item.get("item_name") or ""),
                url_name=str(item.get("url_name") or ""),
                thumb_url=str(item.get("thumb_url") or ""),
                icon_url=str(item.get("icon_url") or ""),
                max_rank=int(item.get("max_rank") or 0),
                is_set=bool(item.get("is_set")),
            )
            for item in items
            if item.get("item_name") and item.get("url_name")
        ]
        return catalog or list(self._market_seed_items)

    def _match_market_item(self, query: str, catalog: list[MarketCatalogItem]) -> MarketCatalogItem:
        normalized_query = query.strip().lower()
        if not normalized_query:
            raise LookupError("Bitte einen Market-Suchbegriff eingeben.")

        for item in catalog:
            if normalized_query in {item.item_name.lower(), item.url_name.lower()}:
                return item
        for item in catalog:
            if normalized_query in item.item_name.lower() or normalized_query in item.url_name.lower():
                return item
        return MarketCatalogItem(
            item_name=query.strip(),
            url_name=self._guess_market_slug(query),
            is_set=" set" in query.lower(),
        )

    def _load_market_details(
        self,
        url_name: str,
        force_refresh: bool,
    ) -> tuple[dict[str, Any], str | None, str]:
        cache_key = f"market:details:{url_name}"
        payload, fetched_at = (
            (None, None)
            if force_refresh
            else self.storage.get_cached_json(cache_key, max_age_seconds=21600)
        )
        if payload is not None:
            return payload, fetched_at, "Item-Details aus SQLite-Cache."

        try:
            payload = self.market_client.fetch_item_details(url_name)
            fetched_at = self.storage.set_cached_json(cache_key, payload)
            return payload, fetched_at, "Item-Details live geladen."
        except RuntimeError as error:
            payload, fetched_at = self.storage.get_any_cached_json(cache_key)
            if payload is not None:
                return payload, fetched_at, f"{error} Item-Details aus Cache."
            raise

    def _load_market_orders(
        self,
        url_name: str,
        force_refresh: bool,
    ) -> tuple[dict[str, Any], str | None, str]:
        cache_key = f"market:orders:{url_name}"
        payload, fetched_at = (
            (None, None)
            if force_refresh
            else self.storage.get_cached_json(cache_key, max_age_seconds=60)
        )
        if payload is not None:
            return payload, fetched_at, "Live-Orders aus SQLite-Cache."

        try:
            payload = {"data": self.market_client.fetch_item_orders(url_name)}
            fetched_at = self.storage.set_cached_json(cache_key, payload)
            return payload, fetched_at, "Live-Orders direkt von warframe.market."
        except RuntimeError as error:
            payload, fetched_at = self.storage.get_any_cached_json(cache_key)
            if payload is not None:
                return payload, fetched_at, f"{error} Live-Orders aus Cache."
            raise

    def _load_market_statistics(
        self,
        url_name: str,
        force_refresh: bool,
    ) -> tuple[dict[str, Any], str | None, str]:
        cache_key = f"market:statistics:{url_name}"
        payload, fetched_at = (
            (None, None)
            if force_refresh
            else self.storage.get_cached_json(cache_key, max_age_seconds=300)
        )
        if payload is not None:
            return payload, fetched_at, "Trade-Stats aus SQLite-Cache."

        try:
            payload = self.market_client.fetch_item_statistics(url_name)
            fetched_at = self.storage.set_cached_json(cache_key, payload)
            return payload, fetched_at, "Trade-Stats live geladen."
        except RuntimeError as error:
            payload, fetched_at = self.storage.get_any_cached_json(cache_key)
            if payload is not None:
                return payload, fetched_at, f"{error} Trade-Stats aus Cache."
            return {}, fetched_at, str(error)

    def _empty_market_snapshot(self, item: MarketCatalogItem, status: str) -> MarketSnapshot:
        return MarketSnapshot(
            item_name=item.item_name,
            url_name=item.url_name,
            fetched_at="-",
            status=status,
            image_url=item.icon_url,
            thumb_url=item.thumb_url,
            max_rank=item.max_rank,
            is_set=item.is_set,
            best_price="-",
            ingame_sellers="0",
            metrics=[],
            orders=[],
            recent_trades=[],
            trend_trades=[],
        )

    def _build_market_snapshot(
        self,
        item: MarketCatalogItem,
        payload: dict[str, Any],
        fetched_at: Any,
        status: str,
    ) -> MarketSnapshot:
        details = payload.get("details", {})
        orders_payload = payload.get("orders", {})
        statistics = payload.get("statistics", {})
        primary_item = self._select_primary_market_item(details, item.url_name)
        localized = primary_item.get("i18n", {}).get("en", {})
        item_name = str(localized.get("name") or payload.get("item_name") or item.item_name)
        url_name = str(primary_item.get("slug") or payload.get("url_name") or item.url_name)
        icon_url = self.market_client._asset_url(localized.get("icon"))
        thumb_url = self.market_client._asset_url(localized.get("thumb"))
        max_rank = int(primary_item.get("maxRank") or item.max_rank or 0)
        is_set = bool("set" in primary_item.get("tags", []) or item.is_set)
        live_orders = self._build_live_orders(
            item_name=item_name,
            orders=orders_payload.get("data", []),
            max_rank=max_rank,
            is_set=is_set,
        )
        recent_entries = self._select_market_rows(statistics.get("48hours", []), limit=8)
        trend_entries = self._select_market_rows(statistics.get("90days", []), limit=12)
        best_price = live_orders[0].price if live_orders else "-"
        ingame_sellers = self._format_count_text(len(live_orders))
        if live_orders:
            status = (
                f"{status} | Es werden aktuell nur sichtbare PC-Sell-Orders mit Status "
                "'ingame' gezeigt."
            )
        else:
            status = f"{status} | Aktuell wurden keine passenden 'ingame'-Seller gefunden."

        return MarketSnapshot(
            item_name=item_name,
            url_name=url_name,
            fetched_at=self._friendly_timestamp(fetched_at),
            status=status,
            image_url=icon_url,
            thumb_url=thumb_url,
            max_rank=max_rank,
            is_set=is_set,
            best_price=best_price,
            ingame_sellers=ingame_sellers,
            metrics=self._build_market_metrics(
                item_name=item_name,
                max_rank=max_rank,
                is_set=is_set,
                live_orders=live_orders,
                recent_entries=recent_entries,
                trend_entries=trend_entries,
            ),
            orders=live_orders,
            recent_trades=[self._build_market_trade_stat(entry) for entry in recent_entries],
            trend_trades=[self._build_market_trade_stat(entry) for entry in trend_entries],
        )

    def _build_market_metrics(
        self,
        item_name: str,
        max_rank: int,
        is_set: bool,
        live_orders: list[MarketLiveOrder],
        recent_entries: list[dict[str, Any]],
        trend_entries: list[dict[str, Any]],
    ) -> list[MarketMetric]:
        latest_recent = recent_entries[0] if recent_entries else {}
        latest_trend = trend_entries[0] if trend_entries else {}
        best_order = live_orders[0] if live_orders else None
        return [
            MarketMetric(
                label="Best Ingame Sell",
                value=best_order.price if best_order else "-",
                detail=best_order.seller_name if best_order else "Kein ingame Seller gefunden",
            ),
            MarketMetric(
                label="Ingame Sellers",
                value=self._format_count_text(len(live_orders)),
                detail="Nur sichtbare Sell-Orders mit Status ingame",
            ),
            MarketMetric(
                label="Item Type",
                value="Set" if is_set else "Single Item",
                detail=f"{item_name} | Max Rank: {max_rank}",
            ),
            MarketMetric(
                label="Latest 48h Avg",
                value=self._format_price_text(latest_recent.get("avg_price")),
                detail=self._format_market_detail(latest_recent),
            ),
            MarketMetric(
                label="Latest 90d Avg",
                value=self._format_price_text(latest_trend.get("avg_price")),
                detail=self._format_market_detail(latest_trend),
            ),
            MarketMetric(
                label="48h Moving Avg",
                value=self._format_price_text(latest_recent.get("moving_avg")),
                detail=self._format_market_range(latest_recent),
            ),
        ]

    def _select_primary_market_item(self, details: dict[str, Any], url_name: str) -> dict[str, Any]:
        items = details.get("items", [])
        for item in items:
            if item.get("slug") == url_name:
                return item
        if items:
            return items[0]
        return {"slug": url_name, "i18n": {"en": {"name": url_name.replace("_", " ").title()}}}

    def _build_live_orders(
        self,
        item_name: str,
        orders: list[dict[str, Any]],
        max_rank: int,
        is_set: bool,
    ) -> list[MarketLiveOrder]:
        filtered_orders: list[MarketLiveOrder] = []
        for order in orders:
            if not isinstance(order, dict):
                continue
            user = order.get("user", {})
            if order.get("type") != "sell":
                continue
            if not order.get("visible", False):
                continue
            if str(user.get("status")).lower() != "ingame":
                continue
            if str(user.get("platform")).lower() != "pc":
                continue

            rank_value = order.get("rank")
            display_name = self._format_order_item_name(
                item_name=item_name,
                rank_value=rank_value,
                max_rank=max_rank,
                is_set=is_set,
            )
            platinum = int(order.get("platinum") or 0)
            seller_name = str(user.get("ingameName") or user.get("slug") or "-")
            filtered_orders.append(
                MarketLiveOrder(
                    seller_name=seller_name,
                    price=f"{platinum}p",
                    quantity=self._format_count_text(order.get("quantity") or 1),
                    rank=self._format_order_rank(rank_value, max_rank=max_rank, is_set=is_set),
                    updated_at=self._friendly_timestamp(order.get("updatedAt")),
                    reputation=str(user.get("reputation") or 0),
                    whisper_text=(
                        f'/w {seller_name} Hi! I want to buy: "{display_name}" '
                        f"for {platinum} platinum."
                    ),
                    platform=str(user.get("platform") or "pc").upper(),
                    crossplay="Yes" if user.get("crossplay") else "-",
                    order_id=str(order.get("id") or ""),
                )
            )

        filtered_orders.sort(
            key=lambda entry: (
                self._safe_float(entry.price.replace("p", "")),
                -self._safe_float(entry.reputation),
                entry.seller_name.lower(),
            )
        )
        return filtered_orders[:30]

    def _select_market_rows(self, rows: list[dict[str, Any]], limit: int) -> list[dict[str, Any]]:
        parsed_rows = [
            row
            for row in rows
            if isinstance(row, dict) and row.get("datetime")
        ]
        parsed_rows.sort(
            key=lambda row: (
                self._parse_datetime(row.get("datetime")) or datetime.min.replace(tzinfo=UTC),
                self._safe_float(row.get("volume")),
                self._safe_float(row.get("mod_rank")),
            ),
            reverse=True,
        )
        return parsed_rows[:limit]

    def _build_market_trade_stat(self, row: dict[str, Any]) -> MarketTradeStat:
        return MarketTradeStat(
            captured_at=self._friendly_timestamp(row.get("datetime")),
            rank=self._format_rank(row.get("mod_rank")),
            avg_price=self._format_price_text(row.get("avg_price")),
            median_price=self._format_price_text(row.get("median")),
            volume=self._format_count_text(row.get("volume")),
            price_range=self._format_market_range(row),
            moving_average=self._format_price_text(row.get("moving_avg")),
        )

    def _format_market_detail(self, row: dict[str, Any]) -> str:
        if not row:
            return "-"
        timestamp = self._friendly_timestamp(row.get("datetime"))
        rank = self._format_rank(row.get("mod_rank"))
        return f"{timestamp} | {rank}"

    def _format_market_range(self, row: dict[str, Any]) -> str:
        if not row:
            return "-"
        minimum = self._format_price_text(row.get("min_price"))
        maximum = self._format_price_text(row.get("max_price"))
        if minimum == "-" and maximum == "-":
            return "-"
        return f"{minimum} - {maximum}"

    def _format_order_item_name(
        self,
        item_name: str,
        rank_value: Any,
        max_rank: int,
        is_set: bool,
    ) -> str:
        if is_set or max_rank <= 0 or rank_value in (None, ""):
            return item_name
        return f"{item_name} (Rank {int(rank_value)})"

    def _format_order_rank(self, rank_value: Any, max_rank: int, is_set: bool) -> str:
        if is_set:
            return "Set"
        if max_rank <= 0:
            return "Base"
        if rank_value in (None, ""):
            return "Base"
        return f"Rank {int(rank_value)}"

    def _serialize_vosfor(self, snapshot: VosforSnapshot) -> dict[str, Any]:
        return {
            "imported_at": snapshot.imported_at,
            "workbook_path": snapshot.workbook_path,
            "gambling_entries": [asdict(entry) for entry in snapshot.gambling_entries],
            "top_entries": [asdict(entry) for entry in snapshot.top_entries],
            "baro_deals": [asdict(entry) for entry in snapshot.baro_deals],
        }

    def _deserialize_vosfor(self, payload: dict[str, Any]) -> VosforSnapshot:
        return VosforSnapshot(
            imported_at=str(payload.get("imported_at") or "Aus SQLite geladen"),
            workbook_path=str(payload.get("workbook_path") or ""),
            gambling_entries=[
                VosforSourceEntry(**entry)
                for entry in payload.get("gambling_entries", [])
            ],
            top_entries=[
                VosforTopEntry(**entry)
                for entry in payload.get("top_entries", [])
            ],
            baro_deals=[
                BaroDeal(**entry)
                for entry in payload.get("baro_deals", [])
            ],
        )

    def _window_status(self, activation: Any, expiry: Any) -> str:
        activation_time = self._parse_datetime(activation)
        expiry_time = self._parse_datetime(expiry)
        now = datetime.now(UTC)

        if activation_time and activation_time > now:
            return f"Starts in {self._format_duration(activation_time - now)}"
        if expiry_time and expiry_time > now:
            return f"Ends in {self._format_duration(expiry_time - now)}"
        return "-"

    def _remaining_from_expiry(self, expiry: Any) -> str:
        expiry_time = self._parse_datetime(expiry)
        if expiry_time is None:
            return "-"
        delta = expiry_time - datetime.now(UTC)
        if delta.total_seconds() <= 0:
            return "expired"
        return self._format_duration(delta)

    def _friendly_timestamp(self, timestamp: Any) -> str:
        parsed = self._parse_datetime(timestamp)
        if parsed is None:
            return "-"
        return parsed.astimezone().strftime("%d.%m.%Y %H:%M")

    @staticmethod
    def _parse_datetime(value: Any) -> datetime | None:
        if not value or not isinstance(value, str):
            return None
        normalized = value.replace("Z", "+00:00")
        try:
            parsed = datetime.fromisoformat(normalized)
        except ValueError:
            return None
        if parsed.tzinfo is None:
            return parsed.replace(tzinfo=UTC)
        return parsed.astimezone(UTC)

    @staticmethod
    def _format_duration(delta) -> str:
        total_seconds = int(delta.total_seconds())
        hours, remainder = divmod(total_seconds, 3600)
        minutes, _ = divmod(remainder, 60)
        if hours > 0:
            return f"{hours}h {minutes}m"
        return f"{minutes}m"

    @classmethod
    def _format_timer_text(cls, value: Any) -> str:
        if value is None:
            return "-"
        text = str(value).strip()
        if not text:
            return "-"
        lowered = text.lower()
        if lowered == "expired":
            return "expired"
        if "h" in text and "m" in text:
            parts = text.split()
            return " ".join(parts[:2])
        if "m" in text and "s" in text:
            return text.split()[0]
        return text

    @staticmethod
    def _format_number(value: Any) -> str:
        if value is None:
            return "-"
        if isinstance(value, (int, float)):
            return f"{value:.5g}"
        return str(value)

    @staticmethod
    def _format_mixed(value: Any) -> str:
        if value is None:
            return "-"
        if isinstance(value, (int, float)):
            text = f"{value:.4f}".rstrip("0").rstrip(".")
            return text or "0"
        return str(value)

    @staticmethod
    def _safe_float(value: Any) -> float:
        try:
            return float(str(value).replace(",", "."))
        except ValueError:
            return 0.0

    @classmethod
    def _format_price_text(cls, value: Any) -> str:
        if value in (None, ""):
            return "-"
        amount = cls._safe_float(value)
        if amount == 0.0 and str(value).strip() not in {"0", "0.0"}:
            return "-"
        if float(amount).is_integer():
            return f"{int(amount)}p"
        return f"{amount:.1f}p"

    @staticmethod
    def _format_count_text(value: Any) -> str:
        if value in (None, ""):
            return "-"
        try:
            amount = int(float(str(value).replace(",", "")))
        except ValueError:
            return str(value)
        return f"{amount:,}"

    @staticmethod
    def _format_rank(value: Any) -> str:
        if value in (None, "", -1):
            return "Base"
        try:
            return f"R{int(value)}"
        except ValueError:
            return str(value)

    @staticmethod
    def _guess_market_slug(query: str) -> str:
        text = query.strip().lower().replace("&", " and ").replace("'", "")
        slug = []
        previous_was_separator = False
        for character in text:
            if character.isalnum():
                slug.append(character)
                previous_was_separator = False
            elif not previous_was_separator:
                slug.append("_")
                previous_was_separator = True
        return "".join(slug).strip("_")

    @staticmethod
    def _closest_key(query: str, options: list[str] | tuple[str, ...] | set[str]) -> str:
        normalized_query = query.strip().lower()
        if not normalized_query:
            raise LookupError("Bitte einen Suchbegriff eingeben.")

        candidates = sorted(options)
        for option in candidates:
            if normalized_query == option.lower():
                return option
        for option in candidates:
            if normalized_query in option.lower():
                return option
        raise LookupError(f"Kein passender Eintrag fuer '{query}' gefunden.")
